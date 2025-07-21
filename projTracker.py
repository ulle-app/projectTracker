import streamlit as st
from openpyxl import load_workbook, Workbook
import json
import os

PROJECT_FILE = "projects.xlsx"
USER_CREDENTIALS_FILE = "user_credentials.json"
DEFAULT_PASSWORD = "password"

if not os.path.exists(USER_CREDENTIALS_FILE):
    initial_credentials = {"admin": "admin123"}
    with open(USER_CREDENTIALS_FILE, 'w') as f:
        json.dump(initial_credentials, f)

with open(USER_CREDENTIALS_FILE, 'r') as f:
    USER_CREDENTIALS = json.load(f)

def save_user_credentials():
    with open(USER_CREDENTIALS_FILE, 'w') as f:
        json.dump(USER_CREDENTIALS, f)

def initialize_excel():
    if not os.path.exists(PROJECT_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Project", "Status", "Progress (%)", "Team Members", "Subtasks"])
        wb.save(PROJECT_FILE)

def load_projects():
    wb = load_workbook(PROJECT_FILE)
    ws = wb.active
    projects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        padded_row = list(row) + [None] * (5 - len(row))
        project = {
            "Project": padded_row[0],
            "Status": padded_row[1],
            "Progress": padded_row[2],
            "TeamMembers": padded_row[3] or "",
            "Subtasks": json.loads(padded_row[4]) if padded_row[4] else []
        }
        projects.append(project)
    return projects

def save_projects_to_excel(projects):
    wb = Workbook()
    ws = wb.active
    ws.append(["Project", "Status", "Progress (%)", "Team Members", "Subtasks"])
    for proj in projects:
        ws.append([
            proj["Project"],
            proj["Status"],
            proj["Progress"],
            proj["TeamMembers"],
            json.dumps(proj["Subtasks"])
        ])
    wb.save(PROJECT_FILE)

initialize_excel()
projects = load_projects()

st.set_page_config(page_title="Project Tracker", page_icon="📈", layout="wide")

if "username" not in st.session_state:
    st.session_state.username = None

if st.session_state.username is None:
    st.title("🔐 Login")
    col1, col2 = st.columns(2)
    with col1:
        username = st.text_input("User")
    with col2:
        password = st.text_input("Pass", type="password")

    if st.button("Login"):
        if username in USER_CREDENTIALS and USER_CREDENTIALS[username] == password:
            st.session_state.username = username
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("Invalid credentials")

if st.session_state.username:
    st.success(f"Welcome, {st.session_state.username}!")

    is_admin = st.session_state.username == "admin"
    st.sidebar.success(f"Logged in as: {st.session_state.username}")

    if st.sidebar.button("Logout"):
        st.session_state.clear()
        st.rerun()

    st.sidebar.subheader("Change Password")
    new_password = st.sidebar.text_input("New Password", type="password")
    if st.sidebar.button("Update Password"):
        USER_CREDENTIALS[st.session_state.username] = new_password
        save_user_credentials()
        st.sidebar.success("Password updated successfully.")

    st.title("📊 Project Dashboard")

    if is_admin:
        st.sidebar.header("➕ Manage Projects")
        new_project_name = st.sidebar.text_input("New Project")
        if st.sidebar.button("Add Project"):
            if new_project_name.strip() and not any(p["Project"] == new_project_name for p in projects):
                projects.append({
                    "Project": new_project_name,
                    "Status": "Not Started",
                    "Progress": 0,
                    "TeamMembers": "",
                    "Subtasks": []
                })
                save_projects_to_excel(projects)
                st.sidebar.success(f"Project '{new_project_name}' added!")

        selected_project = st.sidebar.selectbox("Select Project", [p["Project"] for p in projects])

        new_status = st.sidebar.selectbox("Update Status", ["Not Started", "In Progress", "Completed"])
        if st.sidebar.button("Update Project Status"):
            for proj in projects:
                if proj["Project"] == selected_project:
                    proj["Status"] = new_status
                    save_projects_to_excel(projects)
                    st.sidebar.success(f"Status of '{selected_project}' updated to '{new_status}'.")

        if st.sidebar.button("Delete Project"):
            projects = [proj for proj in projects if proj["Project"] != selected_project]
            save_projects_to_excel(projects)
            st.sidebar.success(f"Project '{selected_project}' has been deleted.")

        st.sidebar.markdown("---")
        st.sidebar.markdown("### 👥 Manage Members")

        member_table = []
        for member in USER_CREDENTIALS.keys():
            if member != "admin":
                assign = st.sidebar.checkbox(f"Assign {member} to project", key=f"assign_{member}")
                member_username = st.sidebar.text_input(f"Username for {member}", value=member, key=f"user_{member}")
                member_password = st.sidebar.text_input(f"Password for {member}", value=USER_CREDENTIALS.get(member, DEFAULT_PASSWORD), key=f"pass_{member}")

                if st.sidebar.button(f"Update {member}", key=f"update_{member}"):
                    USER_CREDENTIALS[member_username] = member_password
                    for proj in projects:
                        if proj["Project"] == selected_project:
                            members = proj["TeamMembers"].split(', ') if proj["TeamMembers"] else []
                            if assign and member_username not in members:
                                members.append(member_username)
                            elif not assign and member_username in members:
                                members.remove(member_username)
                            proj["TeamMembers"] = ', '.join(members)
                    save_user_credentials()
                    save_projects_to_excel(projects)
                    st.sidebar.success(f"{member_username} updated and assignment {'added' if assign else 'removed'} for '{selected_project}'.")

    user_projects = [p for p in projects if is_admin or st.session_state.username in (p['TeamMembers'].split(', ') if p['TeamMembers'] else [])]

    for proj in user_projects:
        st.markdown(f"### 📌 **{proj['Project']}**")
        st.progress(proj['Progress'] / 100)
        st.markdown(f"**Status:** {proj['Status']}")
        st.markdown(f"**Members:** {proj['TeamMembers'] or 'None'}")

        st.markdown("**Subtasks:**")
        if proj['Subtasks']:
            for idx, task in enumerate(proj['Subtasks'], 1):
                st.markdown(f"{idx}. **{task['Member']}**: {task['Description']} - {task['Progress']}% ({task['Status']})")
        else:
            st.info("No subtasks yet.")

        if is_admin or st.session_state.username in (proj['TeamMembers'].split(', ') if proj['TeamMembers'] else []):
            subtask_desc = st.text_input(f"Task Description for {proj['Project']}", key=f"desc_{proj['Project']}")
            subtask_progress = st.slider(f"Progress for {proj['Project']} (%)", 0, 100, key=f"prog_{proj['Project']}")
            subtask_status = st.selectbox("Task Status", ["Planning", "In Progress", "Done"], key=f"stat_{proj['Project']}")

            if st.button(f"Add Task to {proj['Project']}", key=f"addsub_{proj['Project']}"):
                if subtask_desc:
                    proj['Subtasks'].append({
                        "Member": st.session_state.username,
                        "Description": subtask_desc,
                        "Progress": subtask_progress,
                        "Status": subtask_status
                    })
                    total = sum(s['Progress'] for s in proj['Subtasks'])
                    proj['Progress'] = round(total / len(proj['Subtasks']))
                    save_projects_to_excel(projects)
                    st.success("Task added!")
                    # end of file 
